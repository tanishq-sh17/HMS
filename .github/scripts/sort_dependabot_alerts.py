#!/usr/bin/env python3
"""
sort_dependabot_alerts.py
──────────────────────────
Reads a Dependabot alerts Excel file produced by fetch_dependabot_alerts.py,
sorts rows by service name (A→Z) then severity (CRITICAL → HIGH → MEDIUM → LOW),
rewrites the "Alerts" sheet, and updates the "Summary" sheet with per-service
alert counts. Outputs a grouped alerts map for use by w1-jira-manager.

Requirements:
    pip install openpyxl

Usage:
    python sort_dependabot_alerts.py <excel_file>
    python sort_dependabot_alerts.py              # auto-picks latest dependabot_alerts_*.xlsx
"""

import sys
import json
import glob
import os
import subprocess

def _ensure_packages(*packages: str) -> None:
    """Install packages only if they are not already importable."""
    missing = []
    for pkg in packages:
        import importlib
        try:
            importlib.import_module(pkg.split("[")[0].replace("-", "_"))
        except ModuleNotFoundError:
            missing.append(pkg)
    if missing:
        print(f"[INFO] Installing missing packages: {', '.join(missing)}")
        subprocess.check_call([sys.executable, "-m", "pip", "install", *missing])

_ensure_packages("openpyxl")

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ── CONFIG ────────────────────────────────────────────────────────────────────
SEVERITY_ORDER = {"CRITICAL": 0, "HIGH": 1, "MEDIUM": 2, "LOW": 3}
SEVERITY_COLORS = {
    "CRITICAL": "FF4C4C",
    "HIGH":     "FF944C",
    "MEDIUM":   "FFD700",
    "LOW":      "90EE90",
}

COLUMNS = [
    "Service", "Repo", "Alert #", "Severity", "CVE ID",
    "Package", "Vulnerable Range", "Safe Version",
    "Manifest", "Scope", "Summary", "Alert URL",
    "Jira Key", "Jira Status",
]

# Map column header → dict key (matches fetch script's alert dict)
COLUMN_KEYS = [
    "service", "repo", "alert_number", "severity", "cve_id",
    "package", "vulnerable_range", "safe_version",
    "manifest", "scope", "summary", "alert_url",
    "jira_key", "jira_status",
]


# ── FILE RESOLUTION ───────────────────────────────────────────────────────────
def resolve_input_file(arg: str | None) -> str:
    if arg:
        if not os.path.exists(arg):
            print(f"ERROR: File not found: {arg}")
            sys.exit(1)
        return arg

    # Auto-pick the most recently created matching file
    matches = sorted(glob.glob("dependabot_alerts_*.xlsx"), reverse=True)
    if not matches:
        print("ERROR: No dependabot_alerts_*.xlsx file found in current directory.")
        print("       Pass the file path explicitly: python sort_dependabot_alerts.py <file>")
        sys.exit(1)

    chosen = matches[0]
    print(f"[INFO] Auto-selected: {chosen}")
    return chosen


# ── READ ──────────────────────────────────────────────────────────────────────
def read_alerts(ws) -> list[dict]:
    """Read all data rows from the Alerts sheet into a list of dicts."""
    alerts = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):   # skip fully empty rows
            continue
        alert = {key: (row[i] if i < len(row) else "") for i, key in enumerate(COLUMN_KEYS)}
        # Normalise severity to uppercase for consistent sorting
        if alert["severity"]:
            alert["severity"] = str(alert["severity"]).upper()
        alerts.append(alert)
    return alerts


# ── SORT ──────────────────────────────────────────────────────────────────────
def sort_alerts(alerts: list[dict]) -> list[dict]:
    """Sort by service name (A→Z), then severity (CRITICAL first)."""
    return sorted(
        alerts,
        key=lambda a: (
            str(a["service"]).lower(),
            SEVERITY_ORDER.get(str(a["severity"]).upper(), 99),
        )
    )


# ── WRITE ALERTS SHEET ────────────────────────────────────────────────────────
def rewrite_alerts_sheet(ws, alerts: list[dict]) -> None:
    """Overwrite the Alerts sheet with sorted rows, keeping header at row 1."""
    # Clear data rows (keep header)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.value = None
            cell.fill = PatternFill()  # reset fill

    for row_i, a in enumerate(alerts, 2):
        severity_key = str(a["severity"]).upper()
        color = SEVERITY_COLORS.get(severity_key, "FFFFFF")
        fill  = PatternFill("solid", fgColor=color)
        values = [a.get(key, "") for key in COLUMN_KEYS]

        for col_i, val in enumerate(values, 1):
            c = ws.cell(row=row_i, column=col_i, value=val)
            c.fill = fill
            c.alignment = Alignment(wrap_text=True, vertical="top")


# ── WRITE SUMMARY SHEET ───────────────────────────────────────────────────────
def rewrite_summary_sheet(wb, alerts: list[dict]) -> None:
    """Replace the Summary sheet with per-service alert counts."""
    if "Summary" in wb.sheetnames:
        del wb["Summary"]

    ws2 = wb.create_sheet("Summary")

    # Header row
    header = ["Service", "CRITICAL", "HIGH", "MEDIUM", "LOW", "Total"]
    ws2.append(header)
    for cell in ws2[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2C3E50")
        cell.alignment = Alignment(horizontal="center")

    # Aggregate counts per service
    service_counts: dict[str, dict] = {}
    for a in alerts:
        svc = str(a["service"])
        sev = str(a["severity"]).upper()
        if svc not in service_counts:
            service_counts[svc] = {"CRITICAL": 0, "HIGH": 0, "MEDIUM": 0, "LOW": 0}
        if sev in service_counts[svc]:
            service_counts[svc][sev] += 1

    for svc in sorted(service_counts):
        counts = service_counts[svc]
        total  = sum(counts.values())
        ws2.append([svc, counts["CRITICAL"], counts["HIGH"], counts["MEDIUM"], counts["LOW"], total])

    # Column widths
    ws2.column_dimensions["A"].width = 25
    for col_letter in ["B", "C", "D", "E", "F"]:
        ws2.column_dimensions[col_letter].width = 12


# ── GROUP ─────────────────────────────────────────────────────────────────────
def group_by_service(alerts: list[dict]) -> dict[str, list[dict]]:
    """Return {service_name: [alerts sorted by severity]}."""
    grouped: dict[str, list[dict]] = {}
    for a in alerts:
        svc = str(a["service"])
        grouped.setdefault(svc, []).append(a)
    return grouped


# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    print("=" * 60)
    print("  Dependabot Alert Sorter & Filter")
    print("=" * 60)

    input_file = resolve_input_file(sys.argv[1] if len(sys.argv) > 1 else None)

    print(f"\n[READ]  Loading: {input_file}")
    wb = load_workbook(input_file)

    if "Alerts" not in wb.sheetnames:
        print("ERROR: 'Alerts' sheet not found in the workbook.")
        sys.exit(1)

    ws = wb["Alerts"]
    alerts = read_alerts(ws)
    print(f"   -> {len(alerts)} data rows read")

    if not alerts:
        print("\n[OK] No alert rows to sort.")
        sys.exit(0)

    # Sort
    alerts = sort_alerts(alerts)
    print(f"\n[SORT]  Sorted {len(alerts)} rows — service A→Z, then CRITICAL→HIGH→MEDIUM→LOW")

    # Rewrite sheets
    rewrite_alerts_sheet(ws, alerts)
    rewrite_summary_sheet(wb, alerts)

    wb.save(input_file)
    print(f"[SAVE]  Updated: {input_file}")

    # Group for downstream agent
    grouped = group_by_service(alerts)
    services = list(grouped.keys())
    total_per_service = {svc: len(items) for svc, items in grouped.items()}

    print("\n[SUMMARY] Per-service alert counts:")
    for svc, count in total_per_service.items():
        sev_counts = {"CRITICAL": 0, "HIGH": 0, "MEDIUM": 0, "LOW": 0}
        for a in grouped[svc]:
            sev = str(a["severity"]).upper()
            if sev in sev_counts:
                sev_counts[sev] += 1
        parts = ", ".join(f"{k}: {v}" for k, v in sev_counts.items() if v > 0)
        print(f"   {svc:30s} total={count}  ({parts})")

    print(f"\n[OUTPUT] Services found  : {services}")
    print(f"[OUTPUT] Excel file      : {input_file}")
    print(f"[OUTPUT] Grouped map     : (pass to w1-jira-manager)\n")

    # Emit grouped map as JSON for programmatic consumption
    grouped_output = {
        "excel_file": input_file,
        "services": services,
        "total_per_service": total_per_service,
        "grouped_alerts": grouped,
    }
    output_json = input_file.replace(".xlsx", "_grouped.json")
    with open(output_json, "w", encoding="utf-8") as f:
        json.dump(grouped_output, f, indent=2, default=str)
    print(f"[OUTPUT] Grouped JSON    : {output_json}")
    print("         Hand both files to Sub-Agent 3 (Jira Manager).")


if __name__ == "__main__":
    main()
