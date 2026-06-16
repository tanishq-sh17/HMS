#!/usr/bin/env python3
"""
fetch_dependabot_alerts.py
───────────────────────────
Fetches open Dependabot alerts from GitHub repos and exports to a
color-coded Excel file sorted by service name and severity.

Requirements:
    pip install requests openpyxl

Environment Variables:
    GITHUB_TOKEN  — GitHub PAT with security_events + repo scope

Usage:
    python fetch_dependabot_alerts.py
"""

import os
import sys
import requests
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

# Load .env from repo root (two levels up from .github/scripts/)
load_dotenv(dotenv_path=os.path.join(os.path.dirname(__file__), "..", "..", ".env"))

# ── CONFIG ────────────────────────────────────────────────────────────────────
GITHUB_TOKEN = os.getenv("GITHUB_TOKEN", "")

REPOS = [
    "tanishq-sh17/HMS",
    # "your-org/service-2",
    # "your-org/service-3",
]

ECOSYSTEM_FILTER = "maven"   # None = all ecosystems

OUTPUT_FILE = f"dependabot_alerts_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

SEVERITY_ORDER = {"critical": 0, "high": 1, "medium": 2, "low": 3}
SEVERITY_COLORS = {
    "critical": "FF4C4C",
    "high":     "FF944C",
    "medium":   "FFD700",
    "low":      "90EE90",
}

COLUMNS = [
    "Service", "Repo", "Alert #", "Severity", "GHSA ID", "CVE ID",
    "Package", "Vulnerable Range", "Safe Version",
    "Manifest", "Scope", "Summary", "Alert URL",
    "Jira Key", "Jira Status"   # filled later by Sub-Agent 3
]

# ── FETCH ─────────────────────────────────────────────────────────────────────
def fetch_alerts(repo: str) -> list[dict]:
    if not GITHUB_TOKEN:
        print("ERROR: GITHUB_TOKEN environment variable not set.")
        sys.exit(1)

    headers = {
        "Authorization": f"Bearer {GITHUB_TOKEN}",
        "Accept": "application/vnd.github+json",
        "X-GitHub-Api-Version": "2022-11-28",
    }
    alerts = []
    # Dependabot API uses cursor-based pagination (Link header), not page param
    url = f"https://api.github.com/repos/{repo}/dependabot/alerts"
    params = {"state": "open", "per_page": 100}

    while url:
        resp = requests.get(url, headers=headers, params=params)
        if resp.status_code == 403:
            print(f"  [WARN] Access denied for {repo} -- check token permissions (needs security_events scope).")
            break
        if resp.status_code == 404:
            print(f"  [WARN] Repo not found or Dependabot not enabled: {repo}")
            break
        resp.raise_for_status()

        batch = resp.json()
        if not batch:
            break

        for a in batch:
            eco = a["dependency"]["package"]["ecosystem"].lower()
            if ECOSYSTEM_FILTER and eco != ECOSYSTEM_FILTER:
                continue

            adv   = a.get("security_advisory", {})
            vuln  = a.get("security_vulnerability", {})
            patch = vuln.get("first_patched_version")

            alerts.append({
                "service":          repo.split("/")[-1],
                "repo":             repo,
                "alert_number":     a["number"],
                "severity":         adv.get("severity", "unknown").lower(),
                "ghsa_id":          adv.get("ghsa_id", "N/A"),
                "cve_id":           adv.get("cve_id", "N/A"),
                "package":          a["dependency"]["package"]["name"],
                "vulnerable_range": vuln.get("vulnerable_version_range", ""),
                "safe_version":     patch["identifier"] if patch else "No patch available",
                "manifest":         a["dependency"].get("manifest_path", "pom.xml"),
                "scope":            a["dependency"].get("scope", ""),
                "summary":          adv.get("summary", ""),
                "alert_url":        a.get("html_url", ""),
                "jira_key":         "",   # populated by Sub-Agent 3
                "jira_status":      "",   # populated by Sub-Agent 3
            })

        # Follow cursor-based next page from Link header
        link_header = resp.headers.get("Link", "")
        next_url = None
        for part in link_header.split(","):
            if 'rel="next"' in part:
                next_url = part.split(";")[0].strip().strip("<>")
                break
        url = next_url
        params = {}  # params are embedded in the next URL

    return alerts


# ── SORT ──────────────────────────────────────────────────────────────────────
def sort_alerts(alerts: list[dict]) -> list[dict]:
    """Sort by service name (A→Z), then severity (CRITICAL first) within each service."""
    return sorted(
        alerts,
        key=lambda a: (a["service"].lower(), SEVERITY_ORDER.get(a["severity"], 99))
    )


# ── EXPORT ────────────────────────────────────────────────────────────────────
def export_excel(alerts: list[dict], filename: str) -> str:
    wb = Workbook()

    # ── Alerts sheet ──────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Alerts"

    header_fill = PatternFill("solid", fgColor="2C3E50")
    header_font = Font(color="FFFFFF", bold=True)

    for col_i, col_name in enumerate(COLUMNS, 1):
        c = ws.cell(row=1, column=col_i, value=col_name)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    for row_i, a in enumerate(alerts, 2):
        color = SEVERITY_COLORS.get(a["severity"], "FFFFFF")
        fill  = PatternFill("solid", fgColor=color)
        values = [
            a["service"], a["repo"], a["alert_number"],
            a["severity"].upper(), a.get("ghsa_id", "N/A"), a["cve_id"], a["package"],
            a["vulnerable_range"], a["safe_version"],
            a["manifest"], a["scope"], a["summary"],
            a["alert_url"], a["jira_key"], a["jira_status"],
        ]
        for col_i, val in enumerate(values, 1):
            c = ws.cell(row=row_i, column=col_i, value=val)
            c.fill = fill
            c.alignment = Alignment(wrap_text=True, vertical="top")

    for col_i in range(1, len(COLUMNS) + 1):
        ws.column_dimensions[get_column_letter(col_i)].width = 22
    ws.column_dimensions["L"].width = 50  # Summary
    ws.column_dimensions["M"].width = 60  # Alert URL
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.append(["Severity", "Count"])
    for cell in ws2["1:1"]:
        cell.font = Font(bold=True)

    counts = {"CRITICAL": 0, "HIGH": 0, "MEDIUM": 0, "LOW": 0}
    for a in alerts:
        key = a["severity"].upper()
        if key in counts:
            counts[key] += 1

    for sev, cnt in counts.items():
        row = ws2.append([sev, cnt])
        color = SEVERITY_COLORS.get(sev.lower(), "FFFFFF")
        for cell in ws2[ws2.max_row]:
            cell.fill = PatternFill("solid", fgColor=color)

    ws2.column_dimensions["A"].width = 15
    ws2.column_dimensions["B"].width = 10

    wb.save(filename)
    print(f"\n[OK] Excel saved -> {filename}")
    return filename


# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    print("=" * 60)
    print("  Dependabot Alert Fetcher")
    print("=" * 60)

    all_alerts = []
    for repo in REPOS:
        print(f"\n[FETCH] Fetching: {repo}")
        alerts = fetch_alerts(repo)
        print(f"   -> {len(alerts)} open alerts")
        all_alerts.extend(alerts)

    if not all_alerts:
        print("\n[OK] No open Dependabot alerts found.")
        sys.exit(0)

    all_alerts = sort_alerts(all_alerts)

    counts = {}
    for a in all_alerts:
        counts[a["severity"]] = counts.get(a["severity"], 0) + 1

    print(f"\n[SUMMARY] Total: {len(all_alerts)} alerts")
    for sev in ["critical", "high", "medium", "low"]:
        if sev in counts:
            print(f"   {sev.upper()}: {counts[sev]}")

    export_excel(all_alerts, OUTPUT_FILE)
    print(f"\n[OUTPUT] File: {OUTPUT_FILE}")
    print("   Hand this file path to Sub-Agent 2 (Sorter & Filter).")


if __name__ == "__main__":
    main()
